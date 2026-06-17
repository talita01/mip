"""
Gera a MIP estimada do Brasil (68 setores) para anos SEM TRU nível 68 (2022+),
projetando a matriz do ano-base por GRAS com margens da TRU nível 12.

ATENÇÃO — camada de confiabilidade INFERIOR à das matrizes por TRU 68: a estrutura
intersetorial fina é herdada do ano-base e apenas balanceada aos agregados do nível 12.
Validação retroativa (validar_ras.py): desvio médio dos multiplicadores ~2,8-3,0%
(máx ~13-14%, concentrado em setores de choque de preço). Não usar para análise de
mudança estrutural (SDA): a mudança tecnológica fina é, por construção, quase nula.

Saída: dados/estimadas/MIP-BR-mipcore-GRAS-68S-{ano}.xlsx

Uso:  source ~/.venvs/fgv-mip/bin/activate && python gerar_matriz_gras.py [alvo] [base]
      (padrão: alvo 2022, base = último ano com TRU 68 = 2021)
"""
import sys
import warnings
import numpy as np
import openpyxl
import mipcore as mip
from gerar_matrizes_estimadas import _aba_matriz, nomes_setores, SAIDA


def gerar(ano_alvo, ano_base):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        est = mip.gras.estimar(ano_alvo, ano_base)
        mbase = mip.leontief.matrizes(mip.tru.carregar(ano_base))

    # double checks
    rA = np.corrcoef(est["A"].ravel(), mbase["A"].ravel())[0, 1]
    assert rA > 0.99, f"continuidade r(A) {rA}"
    vbp12 = mip.gras.n12_agregados(ano_alvo)["vbp_ativ"].sum()
    desv_vbp = abs(est["g"].sum() / vbp12 - 1)
    assert desv_vbp < 1e-6, f"VBP total difere do nível 12: {desv_vbp}"
    L_chk = np.linalg.inv(np.eye(68) - est["A"])
    assert np.abs(L_chk - est["L"]).max() < 1e-10

    cods = [c.strip() for c in est["atividades"]]
    noms = nomes_setores()
    mp = mip.multiplicadores.producao_tipo1(est["L"])
    tras, frente = mip.multiplicadores.rasmussen_hirschman(est["L"])
    f = est["g"] - est["Z"].sum(1)

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ref = wb.create_sheet("Referência")
    for i, txt in enumerate([
        f"Matriz de Insumo-Produto PROJETADA para o Brasil {ano_alvo} — 68 setores (atividade × atividade)",
        f"MÉTODO: projeção GRAS da matriz de {ano_base} (estimada das TRU 68 oficiais, método híbrido)",
        "com margens-alvo derivadas da TRU nível 12 (CI por bloco nas linhas/colunas, VBP por bloco",
        "para a produção). Algoritmo GRAS na exposição corrigida de LENZEN; WOOD; GALLEGO (2007),",
        "eqs. 6-9, sobre JUNIUS; OOSTERHAVEN (2003). Implementação: mipcore.gras.",
        "",
        "CAMADA DE CONFIABILIDADE INFERIOR à das matrizes por TRU 68: a estrutura fina é herdada",
        f"de {ano_base} e balanceada aos agregados do nível 12. Validação retroativa (backcast 2020/",
        "2019 -> 2021): desvio médio dos multiplicadores tipo I de 2,8-3,0% (máx ~13-14%, concentrado",
        "em setores de choque de preço: carvão, siderurgia, transporte aéreo, eletrônicos).",
        "NÃO usar para análise de mudança estrutural (SDA).",
        f"Controles desta geração: r(A vs {ano_base}) = {rA:.4f}; VBP total = nível 12 (exato);",
        f"fator de fechamento sum(u)=sum(v) = {est['fator_fechamento']:.4f}.",
        "",
        "VAB por aproximação declarada (VAB do ano-base × crescimento do VBP do bloco nível 12).",
        "Valores em R$ milhões correntes (Z) e adimensionais (A, L). Demanda final por resíduo.",
        "Citar: IBGE (TRU níveis 68 e 12); IBGE (2018) MIP Brasil 2015; GUILHOTO; SESSO FILHO",
        "(2005; 2010); LENZEN; WOOD; GALLEGO (2007) Economic Systems Research 19(4) p. 461-465.",
        f"Gerado em {__import__('datetime').date.today().isoformat()} por mipcore v{mip.__version__}.",
    ], start=1):
        ref.cell(i, 1, txt)

    _aba_matriz(wb, "Usos SxS", est["Z"], cods, noms)
    _aba_matriz(wb, "Mat A Coef Tec", est["A"], cods, noms)
    _aba_matriz(wb, "Inv Leontief", est["L"], cods, noms)

    ws = wb.create_sheet("Indicadores")
    cab = ["CÓDIGO", "DESCRIÇÃO", "VBP (g)", "Demanda final (resíduo)", "VAB (aprox.)",
           "Multiplicador produção tipo I", "R-H para trás (U_j)", "R-H para frente (U_i)"]
    for j, c in enumerate(cab):
        ws.cell(1, 1 + j, c)
    for i, c in enumerate(cods):
        vals = [c, noms.get(c, ""), float(est["g"][i]), float(f[i]), float(est["VAB"][i]),
                float(mp[i]), float(tras[i]), float(frente[i])]
        for j, v in enumerate(vals):
            ws.cell(2 + i, 1 + j, v)

    SAIDA.mkdir(exist_ok=True)
    arq = SAIDA / f"MIP-BR-mipcore-GRAS-68S-{ano_alvo}.xlsx"
    wb.save(arq)
    return arq, rA


if __name__ == "__main__":
    args = [int(a) for a in sys.argv[1:]]
    alvo = args[0] if args else 2022
    base = args[1] if len(args) > 1 else 2021
    arq, rA = gerar(alvo, base)
    print(f"{alvo}: {arq.name} gerado (base {base}) | r(A vs base)={rA:.4f}")
