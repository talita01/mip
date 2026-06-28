"""mipcore.regional — sistema inter-regional de insumo-produto (IIOAS, 27 UFs × 68 setores).

Carrega a Matriz Interestadual de Insumo-Produto (Haddad/NEREUS) e expõe o ferramental
analítico por estado e inter-regional: inversa de Leontief, multiplicadores de produção,
índices de Rasmussen-Hirschman, autossuficiência por UF, extração hipotética de região e
matriz de multiplicadores-produto (MPM).

Anos disponíveis (validados — ver Manual-Metodologico-Insumo-Produto.md, §6.3):
  2011 — HADDAD; GONÇALVES JUNIOR; NASCIMENTO (2017), arquivo na raiz da pasta, com A, I-A
         e B=(I-A)^-1 prontas. Lê-se A e recalcula-se L=inv(I-A) (coincide com o B do
         arquivo a ~1e-15).
  2019 — HADDAD; ARAÚJO; ROCHA; VALE (2025), arquivo `IIOAS_BRUF_2019.xlsx` (dentro do .rar
         na raiz). Traz fluxos (aba "MIIP SS"), não A/L prontas: deriva-se A = Z·diag(1/VBP),
         com Z = bloco intermediário (linhas 7-1842, colunas 5-1840) e VBP = "Valor da
         produção" (linha 1856). Identidade Z+importação+impostos+VAB = VBP confere a 0%
         (conferido em 17 jun. 2026). As contas socioambientais do arquivo (CO2, água,
         energia — linhas 1861-1863) NÃO são carregadas aqui: por diretriz, conta-satélite
         fica fora desta pasta.

Estrutura do índice (igual nos dois anos): ordenado por região (region-major). O índice
plano de (UF r = 0..26, setor s = 0..67) é r*68 + s.

Arquivos grandes (130/115 MB) são lidos uma vez e cacheados em ~/.cache/mipcore/ (fora do
OneDrive). O .xlsx de 2019 deve estar no cache ou em /tmp/iioas2019/ (extrair o .rar).
"""
import json
import re
from pathlib import Path

import numpy as np
import openpyxl

from . import tru, multiplicadores

NR, NS = 27, 68
N = NR * NS  # 1836

_ROOT = tru.DADOS.parents[3]
_CACHE = Path.home() / ".cache" / "mipcore"
_ARQ_2011 = _ROOT / "administrador,+IIOAS_Brasil_RBERU_2017.xlsx"
_RAR_2019 = _ROOT / "IIOAS_BRUF_2019 (1).rar"
_XLSX_2019 = [_CACHE / "IIOAS_BRUF_2019.xlsx", Path("/tmp/iioas2019/IIOAS_BRUF_2019.xlsx")]


# ---------------- leitura 2011 (A pronta na aba "A") ----------------
def _ler_A_2011(arq):
    """Bloco numérico 1836×1836 da aba A (canto em linha 4, coluna 4)."""
    wb = openpyxl.load_workbook(str(arq), read_only=True, data_only=True)
    ws = wb["A"]
    A = np.empty((N, N))
    r = 0
    for row in ws.iter_rows(min_row=4, max_row=3 + N, min_col=4, max_col=3 + N,
                            values_only=True):
        A[r, :] = [(x if isinstance(x, (int, float)) else 0.0) for x in row]
        r += 1
    wb.close()
    assert r == N, f"esperadas {N} linhas, lidas {r}"
    return A


def _ler_rotulos_2011(arq):
    wb = openpyxl.load_workbook(str(arq), read_only=True, data_only=True)
    regioes = []
    for row in wb["Regions"].iter_rows(values_only=True):
        v = [c for c in row if c not in (None, "")]
        if len(v) >= 3 and str(v[0]).startswith("R") and ":" in str(v[1]):
            regioes.append((str(v[0]), str(v[1]).split(":")[1], str(v[2])))
    setores = []
    for row in wb["Sectors"].iter_rows(values_only=True):
        v = [c for c in row if c not in (None, "")]
        if len(v) >= 3 and isinstance(v[0], (int, float)):
            setores.append((int(v[0]), str(v[1]), str(v[2])))
    wb.close()
    assert len(regioes) == NR and len(setores) == NS, \
        f"rótulos 2011: {len(regioes)} regiões, {len(setores)} setores"
    return regioes, setores


# ---------------- leitura 2019 (derivar A dos fluxos da aba "MIIP SS") ----------------
def _fonte_2019():
    for c in _XLSX_2019:
        if c.exists():
            return c
    raise FileNotFoundError(
        "IIOAS_BRUF_2019.xlsx não encontrado. Extraia o .rar para o cache (fora do OneDrive):\n"
        f"  unar -o {_CACHE} \"{_RAR_2019}\"\n"
        "O .xlsx tem ~115 MB.")


def _ler_A_2019(arq):
    """Deriva A da aba 'MIIP SS': Z (linhas 7-1842, colunas 5-1840) / VBP (linha 1856)."""
    wb = openpyxl.load_workbook(str(arq), read_only=True, data_only=True)
    ws = wb["MIIP SS"]
    Z = np.empty((N, N))
    vbp = None
    for r, row in enumerate(ws.iter_rows(min_row=7, max_row=1856, min_col=5, max_col=4 + N,
                                         values_only=True), start=7):
        if r <= 6 + N:                       # 7..1842 = bloco intermediário
            Z[r - 7, :] = [(x if isinstance(x, (int, float)) else 0.0) for x in row]
        elif r == 1856:                      # Valor da produção (VBP)
            vbp = np.array([(x if isinstance(x, (int, float)) else 0.0) for x in row])
    wb.close()
    assert vbp is not None, "VBP (linha 1856) não encontrado"
    # setor sem produção numa UF (VBP=0) fica com coeficientes nulos (0/0 -> 0)
    pos = vbp > 0
    A = np.zeros((N, N))
    A[:, pos] = Z[:, pos] / vbp[pos][None, :]
    return A


def _ler_rotulos_2019(arq):
    wb = openpyxl.load_workbook(str(arq), read_only=True, data_only=True)
    regioes = []
    for row in wb["Regiões"].iter_rows(values_only=True):
        v = [c for c in row if c not in (None, "")]
        if len(v) >= 3 and re.fullmatch(r"R\d+", str(v[0])):
            regioes.append((str(v[0]), str(v[1]), str(v[2])))
    setores = []
    for row in wb["Setores"].iter_rows(values_only=True):
        v = [c for c in row if c not in (None, "")]
        if len(v) >= 2 and re.fullmatch(r"S\d+", str(v[0])):
            setores.append((int(str(v[0])[1:]), str(v[0]), str(v[1])))
    wb.close()
    assert len(regioes) == NR and len(setores) == NS, \
        f"rótulos 2019: {len(regioes)} regiões, {len(setores)} setores"
    return regioes, setores


# ---------------- demanda final ----------------
_CAT_2019 = ["Investimento", "Famílias", "Governo", "ISFLSF", "Exportações"]


def _ler_fd_2019(arq):
    """Demanda final da aba MIIP SS: 5 categorias × 27 regiões (cols 1841-1975) + discrepância
    estatística (col 1976). Retorna (total por (UF,setor) produtor; dict categoria->vetor
    somado sobre as 27 regiões demandantes). Identidade L·total = VBP confere a 0% (§6.3)."""
    wb = openpyxl.load_workbook(str(arq), read_only=True, data_only=True)
    ws = wb["MIIP SS"]
    FD = np.zeros((N, 136))                       # colunas 1841..1976
    for r, row in enumerate(ws.iter_rows(min_row=7, max_row=6 + N, min_col=1841, max_col=1976,
                                         values_only=True), start=7):
        FD[r - 7, :] = [(x if isinstance(x, (int, float)) else 0.0) for x in row]
    wb.close()
    cat = {nome: FD[:, i * NR:(i + 1) * NR].sum(1) for i, nome in enumerate(_CAT_2019)}
    return FD.sum(1), cat                          # total inclui a discrepância (col 1976)


def _ler_vbp_2011(arq):
    """Produção bruta (VBP) da aba IIOS, linha 1917, colunas 4-1839."""
    wb = openpyxl.load_workbook(str(arq), read_only=True, data_only=True)
    ws = wb["IIOS"]
    vbp = None
    for row in ws.iter_rows(min_row=1917, max_row=1917, min_col=4, max_col=3 + N,
                            values_only=True):
        vbp = np.array([(x if isinstance(x, (int, float)) else 0.0) for x in row])
    wb.close()
    assert vbp is not None and (vbp > 0).all(), "VBP 2011 (IIOS linha 1917) inválido"
    return vbp


# contas regionais por (UF,setor), aba MIIP SS, colunas 5-1840 (linha 1-indexada do Excel)
_CONTAS_2019_LINHAS = {"imp_produto": 1846, "vab": 1851, "remun": 1852, "eob": 1853,
                       "imp_prod": 1854, "x": 1856, "emp": 1858}


def _ler_contas_2019(arq):
    """Vetores de contas por (UF,setor) (N,): VBP, VAB, remunerações, EOB, impostos sobre a
    produção, impostos sobre o produto e ocupações. Lidos das linhas de totais da aba MIIP SS."""
    wb = openpyxl.load_workbook(str(arq), read_only=True, data_only=True)
    ws = wb["MIIP SS"]
    alvo = set(_CONTAS_2019_LINHAS.values())
    buf = {}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max(alvo), min_col=5, max_col=4 + N,
                                         values_only=True), start=1):
        if r in alvo:
            buf[r] = np.array([(v if isinstance(v, (int, float)) else 0.0) for v in row])
    wb.close()
    return {k: buf[ln] for k, ln in _CONTAS_2019_LINHAS.items()}


def _carregar_contas(ano, usar_cache):
    """Contas regionais (só 2019; o arquivo de 2011 não traz esse bloco na mesma estrutura)."""
    if ano != 2019:
        return None
    cache = _CACHE / f"iioas_{ano}_contas.npz"
    if usar_cache and cache.exists():
        d = np.load(cache)
        return {k: d[k] for k in d.files}
    c = _ler_contas_2019(_fonte_2019())
    np.savez(cache, **c)
    return c


def _carregar_fd(ano, A, usar_cache):
    """Demanda final total e (para 2019) por categoria, com cache em ~/.cache/mipcore."""
    cache_f = _CACHE / f"iioas_{ano}_fd.npz"
    if usar_cache and cache_f.exists():
        d = np.load(cache_f, allow_pickle=True)
        nomes = list(d["cat_nomes"])
        cat = {n: v for n, v in zip(nomes, d["cat_vals"])} if nomes else None
        return d["total"], cat
    if ano == 2019:
        total, cat = _ler_fd_2019(_fonte_2019())
        np.savez(cache_f, total=total, cat_nomes=np.array(list(cat)),
                 cat_vals=np.array(list(cat.values())))
    else:                                          # 2011: f = (I-A)·VBP (exato)
        vbp = _ler_vbp_2011(_ARQ_2011)
        with np.errstate(all="ignore"):
            total = (np.eye(N) - A) @ vbp
        cat = None
        np.savez(cache_f, total=total, cat_nomes=np.array([]), cat_vals=np.array([]))
    return total, cat


# ---------------- carregamento ----------------
def carregar(ano=2011, usar_cache=True):
    """Carrega o sistema inter-regional: A, L=inv(I-A), rótulos e índices.

    ano: 2011 (A pronta) ou 2019 (A derivada dos fluxos). Retorna dict com ano, A (N×N),
    L (N×N), uf (27 siglas na ordem R1..R27), regioes [(R,UF,nome)], setores [(num,cod,nome)].
    """
    if ano not in (2011, 2019):
        raise NotImplementedError(f"ano {ano} indisponível; use 2011 ou 2019.")
    _CACHE.mkdir(parents=True, exist_ok=True)
    cache_A = _CACHE / f"iioas_{ano}_A.npy"
    cache_rot = _CACHE / f"iioas_{ano}_rotulos.json"

    if usar_cache and cache_A.exists() and cache_rot.exists():
        A = np.load(cache_A)
        rot = json.loads(cache_rot.read_text())
        regioes = [tuple(x) for x in rot["regioes"]]
        setores = [tuple(x) for x in rot["setores"]]
    else:
        if ano == 2011:
            if not _ARQ_2011.exists():
                raise FileNotFoundError(f"matriz IIOAS 2011 não encontrada em {_ARQ_2011}")
            A = _ler_A_2011(_ARQ_2011)
            regioes, setores = _ler_rotulos_2011(_ARQ_2011)
        else:
            arq = _fonte_2019()
            A = _ler_A_2019(arq)
            regioes, setores = _ler_rotulos_2019(arq)
        np.save(cache_A, A)
        cache_rot.write_text(json.dumps({"regioes": regioes, "setores": setores}))

    with np.errstate(all="ignore"):   # warnings de BLAS no matmul são benignos neste ambiente
        L = np.linalg.inv(np.eye(N) - A)
        assert A.sum(0).max() < 1, f"coluna de A com soma >= 1 ({A.sum(0).max():.3f})"
        assert np.abs((np.eye(N) - A) @ L - np.eye(N)).max() < 1e-6, "L não é inv(I-A)"
        assert not (np.isnan(L).any() or np.isinf(L).any()), "L com NaN/Inf"

    f, f_cat = _carregar_fd(ano, A, usar_cache)
    contas = _carregar_contas(ano, usar_cache)

    return {"ano": ano, "A": A, "L": L, "f": f, "f_categorias": f_cat, "contas": contas,
            "uf": [r[1] for r in regioes], "regioes": regioes, "setores": setores}


def produtos(ano=2019):
    """Lista dos 128 produtos (código, descrição) da classificação da matriz interestadual (2019).

    Permite desambiguar atividades que o nível 68 agrega: ex. `P054` "Adubos e fertilizantes"
    fica dentro do setor S21 (químicos), e os produtos de milho `P002`/`P032` ajudam a tratar o
    etanol de milho. Para os coeficientes de uso produto×setor, a aba `MIIP PS` da base traz a
    matriz completa (128 produtos × 68 setores) — leitura ainda não exposta aqui.
    """
    if ano != 2019:
        raise NotImplementedError("lista de produtos disponível apenas para 2019")
    cache = _CACHE / f"iioas_{ano}_produtos.json"
    if cache.exists():
        return [tuple(x) for x in json.loads(cache.read_text())]
    wb = openpyxl.load_workbook(str(_fonte_2019()), read_only=True, data_only=True)
    out = []
    for row in wb["Produtos"].iter_rows(values_only=True):
        v = [c for c in row if c not in (None, "")]
        if len(v) >= 2 and re.fullmatch(r"P\d+", str(v[0])):
            out.append((str(v[0]), str(v[1])))
    wb.close()
    assert len(out) == 128, f"esperados 128 produtos, lidos {len(out)}"
    cache.write_text(json.dumps(out))
    return out


def demanda_final(sys):
    """Demanda final do sistema. Retorna dict com 'total' (vetor N por (UF,setor) produtor) e
    'categorias' (dict categoria->vetor, somado sobre as UFs demandantes; só 2019, None em 2011).

    Validada: L @ total reproduz o VBP (2019 a 0% por leitura das colunas 1841-1976; 2011 por
    construção, pois total = (I-A)·VBP). Composição 2019: famílias ~R$4,0 tri, governo ~1,5,
    investimento ~0,95, exportações ~1,0, ISFLSF ~0,1.
    """
    return {"total": sys["f"], "categorias": sys.get("f_categorias")}


# ---------------- índices e blocos ----------------
def idx(sys, uf, setor=None):
    """Índice(s) plano(s). Sem `setor`: os 68 índices da UF. Com `setor` (cód ou nº): um."""
    r = sys["uf"].index(uf)
    if setor is None:
        return np.arange(r * NS, (r + 1) * NS)
    if isinstance(setor, str):
        cods = [s[1] for s in sys["setores"]]
        s = cods.index(setor)
    else:
        s = int(setor) - 1
    return r * NS + s


def bloco(M, uf_origem_i, uf_destino_i):
    """Bloco 68×68 da UF de origem (linhas) para a UF de destino (colunas), por índice 0..26."""
    ro, rd = uf_origem_i, uf_destino_i
    return M[ro * NS:(ro + 1) * NS, rd * NS:(rd + 1) * NS]


# ---------------- multiplicadores ----------------
def multiplicador_producao(sys):
    """Multiplicador de produção tipo I por (UF, setor). Retorna array (27, 68)."""
    return multiplicadores.producao_tipo1(sys["L"]).reshape(NR, NS)


def rasmussen_hirschman(sys):
    """Índices R-H para trás e para frente por (UF, setor). Retorna (tras, frente), cada (27,68)."""
    tras, frente = multiplicadores.rasmussen_hirschman(sys["L"])
    return tras.reshape(NR, NS), frente.reshape(NR, NS)


def autossuficiencia_uf(sys):
    """Participação dos insumos intra-regionais no total de insumos intermediários, por UF.

    Para cada UF r: soma do bloco A^{rr} (insumos comprados dentro da UF) sobre a soma de
    todos os insumos comprados pela UF (todas as origens). Próximo de 1 = mais autossuficiente.
    """
    A = sys["A"]
    out = {}
    for r, uf in enumerate(sys["uf"]):
        cols = slice(r * NS, (r + 1) * NS)
        intra = A[r * NS:(r + 1) * NS, cols].sum()
        total = A[:, cols].sum()
        out[uf] = float(intra / total) if total > 0 else float("nan")
    return out


# ---------------- extração hipotética de região ----------------
def extracao_regiao(sys, uf, f=None):
    """Extração hipotética de uma UF: zera as linhas e colunas da região em A, recalcula a
    produção total para a mesma demanda final e mede a queda.

    f: vetor de demanda final (N,). Sem f, usa a **demanda final real** do sistema (sys["f"],
    validada para reproduzir o VBP); só cai no choque unitário estrutural se ela faltar.
    Retorna dict com produção total com e sem a região e o impacto absoluto e relativo.
    """
    if f is None:
        f = sys.get("f")
    if f is None:
        nota = "f unitário (estrutural)"
        f = np.ones(N)
    else:
        nota = "demanda final real" if f is sys.get("f") else "f fornecido"
    A = sys["A"].copy()
    r = sys["uf"].index(uf)
    sl = slice(r * NS, (r + 1) * NS)
    A[sl, :] = 0.0
    A[:, sl] = 0.0
    with np.errstate(all="ignore"):   # warnings de BLAS no matmul são benignos
        L_bar = np.linalg.inv(np.eye(N) - A)
        x = sys["L"] @ f
        x_bar = L_bar @ f
    T, T_bar = float(x.sum()), float(x_bar.sum())
    return {"uf": uf, "producao_total": T, "producao_sem_regiao": T_bar,
            "impacto_absoluto": T - T_bar, "impacto_relativo": (T - T_bar) / T, "nota": nota}


# ---------------- matriz de multiplicadores-produto ----------------
def mpm(M):
    """Matriz de multiplicadores-produto de Sonis-Hewings: M_ij = (1/V) R_i C_j (§4.15)."""
    V = M.sum()
    return np.outer(M.sum(1), M.sum(0)) / V


def agregar_uf(M):
    """Agrega uma matriz N×N para 27×27 somando os blocos por par de UFs."""
    return M.reshape(NR, NS, NR, NS).sum(axis=(1, 3))


# ---------------- colapso para 2 regiões (SP × RB) e fechamento Tipo II ----------------
def colapsar_sp_rb(sys, uf="SP"):
    """Colapsa o sistema inter-regional de 27 UFs em 2 regiões: a UF dada (ex. SP) e o Resto
    do Brasil (RB = as 26 demais), preservando os 68 setores. Agrega FLUXOS (Z = A·diag(VBP)) e
    totais — não coeficientes, que não somam linearmente —, recompõe A2 = Z2·diag(1/VBP2),
    L2 = inv(I-A2), demanda final e contas-satélite, e devolve um `sys` de ordem 2·68 = 136 no
    mesmo formato de `carregar`, com NR=2. Requer as contas regionais (VBP), só em 2019.

    Preserva as identidades: soma de coluna de A2 < 1; L2 = inv(I-A2); L2·f2 reproduz o VBP
    agregado; e o VBP total e o da UF alvo são conservados na agregação. As funções genéricas
    de `multiplicadores` operam sobre `sys["L"]` de qualquer ordem; para remodelar use
    `.reshape(sys["NR"], sys["NS"])`.
    """
    if sys.get("contas") is None:
        raise ValueError("colapso SP×RB requer as contas regionais (use carregar(2019))")
    if uf not in sys["uf"]:
        raise ValueError(f"UF {uf!r} não está no sistema")
    x = sys["contas"]["x"]                                    # VBP por (UF, setor), N
    Z = sys["A"] * x[None, :]                                 # fluxos intermediários N×N
    grupo = np.ones(NR, dtype=int)                            # 1 = RB
    grupo[sys["uf"].index(uf)] = 0                            # 0 = UF alvo
    G = 2
    P = np.zeros((G * NS, N))                                 # agrega região→grupo, mantém setor
    for r in range(NR):
        P[grupo[r] * NS + np.arange(NS), r * NS + np.arange(NS)] = 1.0
    Z2 = P @ Z @ P.T                                          # fluxos agregados 136×136
    x2 = P @ x                                                # VBP agregado 136
    pos = x2 > 0
    A2 = np.zeros((G * NS, G * NS))
    with np.errstate(all="ignore"):   # warnings de BLAS no matmul são benignos neste ambiente
        A2[:, pos] = Z2[:, pos] / x2[pos][None, :]
        L2 = np.linalg.inv(np.eye(G * NS) - A2)
        f2, contas2 = P @ sys["f"], {k: P @ v for k, v in sys["contas"].items()}
        f_cat2 = ({k: P @ v for k, v in sys["f_categorias"].items()}
                  if sys.get("f_categorias") else None)
    return {"ano": sys["ano"], "A": A2, "L": L2, "f": f2, "f_categorias": f_cat2,
            "contas": contas2, "uf": [uf, "RB"],
            "regioes": [("R1", uf, uf), ("R2", "RB", "Resto do Brasil")],
            "setores": sys["setores"], "NR": G, "NS": NS}


def fechar_familias_regional(sys, alpha=1.0):
    """Fecha um sistema regional nas famílias (multiplicadores Tipo II), aceitando o sistema
    cheio (27 UFs) ou o colapsado SP×RB. Usa renda w = remunerações/VBP e cesta de consumo =
    categoria 'Famílias' da demanda final. Devolve o dict de `multiplicadores.fechar_familias`
    (Abar, Lbar, rho, n). Requer contas regionais e categorias de demanda final (2019).
    """
    c, fc = sys.get("contas"), sys.get("f_categorias")
    if c is None or not fc or "Famílias" not in fc:
        raise ValueError("fechamento requer contas regionais e a categoria 'Famílias' (2019)")
    x = c["x"]
    w = np.divide(c["remun"], np.where(x > 0, x, 1.0))
    return multiplicadores.fechar_familias(sys["A"], w, fc["Famílias"], alpha=alpha)
