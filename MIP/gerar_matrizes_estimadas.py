"""
Gera as Matrizes de Insumo-Produto estimadas do Brasil (68 setores) a partir das TRU
oficiais do IBGE, com o pipeline mipcore (validado contra Guilhoto 2010-2018 e contra a
MIP oficial 2015 — ver Manual-Metodologico-Insumo-Produto.md, §1.4).

Saída: dados/estimadas/MIP-BR-mipcore-68S-{ano}.xlsx, com abas:
  Referência     — método, validação e citações obrigatórias
  Usos SxS       — fluxos intersetoriais Z (atividade × atividade, nacional, preços básicos)
  Mat A Coef Tec — coeficientes técnicos A = D·Bn (atividade × atividade)
  Inv Leontief   — L = (I-A)^-1
  Indicadores    — g, demanda final (resíduo), VAB, multiplicador de produção tipo I,
                   índices de Rasmussen-Hirschman (para trás / para frente)

Uso:  source ~/.venvs/fgv-mip/bin/activate && python gerar_matrizes_estimadas.py [anos...]
"""
import sys
import warnings
import numpy as np
import openpyxl
import mipcore as mip

SAIDA = mip.tru.DADOS.parent.parent / "estimadas"   # MIP/dados/estimadas/


def nomes_setores():
    """Nomes oficiais das 68 atividades (cabeçalho da TRU, aba producao)."""
    import xlrd
    ws = xlrd.open_workbook(mip.tru.DADOS / "68_tab1_2019.xls").sheet_by_name("producao")
    return {mip.tru._cod(ws.cell_value(3, c), 4):
            str(ws.cell_value(3, c)).split("\n")[1].strip() if "\n" in str(ws.cell_value(3, c))
            else "" for c in range(2, 2 + mip.tru.N_ATIV)}


def _aba_matriz(wb, nome, M, cods, noms):
    ws = wb.create_sheet(nome)
    ws.cell(1, 1, "CÓDIGO"); ws.cell(1, 2, "DESCRIÇÃO")
    for j, c in enumerate(cods):
        ws.cell(1, 3 + j, c)
        ws.cell(2, 3 + j, noms.get(c, ""))
    for i, c in enumerate(cods):
        ws.cell(3 + i, 1, c); ws.cell(3 + i, 2, noms.get(c, ""))
        for j in range(len(cods)):
            ws.cell(3 + i, 3 + j, float(M[i, j]))


def gerar(ano):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        d = mip.tru.carregar(ano)
        ids = mip.tru.checar_identidades(ano)
        m = mip.leontief.matrizes(d)
        chk = mip.leontief.checar(d)
    assert ids["CI+DF=oferta_pc"] < 1e-3 and ids["g=CI_col+VAB(max)"] < 1e-3, f"identidades {ano}"
    assert chk["colunas_A>=1"] == 0 and chk["max|Ly-g|"] < 1e-6, f"controles {ano}"

    cods = [c.strip() for c in m["atividades"]]
    noms = nomes_setores()
    f = m["g"] - m["Z"].sum(1)
    mp = mip.multiplicadores.producao_tipo1(m["L"])
    tras, frente = mip.multiplicadores.rasmussen_hirschman(m["L"])

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ref = wb.create_sheet("Referência")
    for i, txt in enumerate([
        f"Matriz de Insumo-Produto estimada para o Brasil {ano} — 68 setores (atividade × atividade)",
        "Estimação: pipeline mipcore (FGV-Bioeconomia), a partir das TRU oficiais do IBGE (nível 68).",
        "Passagem a preços básicos pelo MÉTODO HÍBRIDO: participações de destino célula a célula das",
        "tabelas oficiais 04-10 da MIP 2015 aplicadas aos totais de margens/impostos/importações do",
        "próprio ano (aba oferta da TRU), fallback alpha_ij de Guilhoto-Sesso (2005), realocação das",
        "margens às linhas de comércio e transporte. Modelo de tecnologia do setor (D·Bn), IBGE (2018).",
        "",
        "Validação (ver Manual-Metodologico-Insumo-Produto.md, §1.4 e §2.4): contra a MIP oficial 2015,",
        "r(L) = 1,0000 e desvio dos multiplicadores de 0,15%; contra as matrizes de Guilhoto 2010-2018,",
        "r(L) ≥ 0,9987 e desvio médio de 1,8%; benchmark fora da amostra (Tabela 03 de 2010) com WMAPE",
        "de 5,5% (vs 8,1% Guilhoto-Sesso puro e 16,4% do método r_int anterior).",
        f"Controles deste ano: max|L·f − g| = {chk['max|Ly-g|']:.2e}; identidades contábeis das TRU = 0.",
        "",
        "Valores em R$ milhões correntes (Z) e adimensionais (A, L). Demanda final por resíduo.",
        "Citar obrigatoriamente: IBGE (2018) Matriz de insumo-produto Brasil 2015; GUILHOTO &",
        "SESSO FILHO (2005; 2010) para a metodologia de estimação com dados preliminares.",
        f"Gerado em {__import__('datetime').date.today().isoformat()} por mipcore v{mip.__version__}.",
    ], start=1):
        ref.cell(i, 1, txt)

    _aba_matriz(wb, "Usos SxS", m["Z"], cods, noms)
    _aba_matriz(wb, "Mat A Coef Tec", m["A"], cods, noms)
    _aba_matriz(wb, "Inv Leontief", m["L"], cods, noms)

    ws = wb.create_sheet("Indicadores")
    cab = ["CÓDIGO", "DESCRIÇÃO", "VBP (g)", "Demanda final (resíduo)", "VAB",
           "Multiplicador produção tipo I", "R-H para trás (U_j)", "R-H para frente (U_i)"]
    for j, c in enumerate(cab):
        ws.cell(1, 1 + j, c)
    for i, c in enumerate(cods):
        vals = [c, noms.get(c, ""), float(m["g"][i]), float(f[i]), float(m["VAB"][i]),
                float(mp[i]), float(tras[i]), float(frente[i])]
        for j, v in enumerate(vals):
            ws.cell(2 + i, 1 + j, v)

    SAIDA.mkdir(exist_ok=True)
    arq = SAIDA / f"MIP-BR-mipcore-68S-{ano}.xlsx"
    wb.save(arq)
    return arq, chk


if __name__ == "__main__":
    anos = [int(a) for a in sys.argv[1:]] or [2019, 2020, 2021]
    for ano in anos:
        arq, chk = gerar(ano)
        print(f"{ano}: {arq.name} gerado | max|Lf-g|={chk['max|Ly-g|']:.1e} | "
              f"max colsum A={chk['max_soma_coluna_A']:.3f}")
